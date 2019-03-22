from openpyxl.styles.fills import PatternFill

class Color:
    def __init__(self, color):
        self.color = color
    def __get__(self, obj, objtype=None):
        return self.color
    def __set__(self, obj, value):
        pass
    def __delete__(self, obj):
        pass

class OpTypeColor:
    CREATE = Color(PatternFill(fill_type="solid", start_color="FF0000", end_color="FF0000"))
    DELETE = Color(PatternFill(fill_type="solid", start_color="00FF00", end_color="00FF00"))
    UPDATE = Color(PatternFill(fill_type="solid", start_color="0000FF", end_color="0000FF"))
    UP_CONTENT = Color(PatternFill(fill_type="solid", start_color="663366", end_color="663366"))
    ENTITY_NOT_EXIST = Color(PatternFill(fill_type="solid", start_color="FFFF00", end_color="FFFF00"))
    

#-----------------------------------------------------------------#

def top():
    def __get__():
        print("__get__")
    print("dd")
    return __get__

class T:
    tp = top
print(T.tp)
a = top()
b = top()
print(a)
print(b)
print(top.__get__)
print(top.__get__)