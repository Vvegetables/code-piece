#coding=utf-8
from functools import wraps
import os
import re
import sys

import jieba
import nltk
from nltk.collocations import BigramCollocationFinder, TrigramCollocationFinder
from nltk.corpus import stopwords, conll2000
from nltk.corpus.reader.plaintext import PlaintextCorpusReader
from nltk.util import trigrams
from openpyxl.chart import pie_chart
from openpyxl.reader.excel import load_workbook
import pymysql


#进行词性标注
def ie_proprocess(document):
    sentences = nltk.sent_tokenize(document)
    print(sentences) #句子分割
    sentences = [nltk.word_tokenize(sent) for sent in sentences]
    print(sentences) #分词器
    sentences = [nltk.pos_tag(sent) for sent in sentences]
    print(sentences) #词性标注
    
    return sentences

#自动获得collocations    
def get_phrase():
    root_dir = r'E:\github_repo\python_basic\pythonbasictest\self_nltk\files'
    wordlists = PlaintextCorpusReader(root_dir,".*")
    x = nltk.Text(wordlists.words("test.txt"))
    print(x)
    
    print(x.collocations())

# def change_stdout(func):
#     @wraps
#     def wrapper(*args,**kwargs):
#         
#         return func(*args,**kwargs)
#     
#     return wrapper

def get_phrase_2(text):
#     text = '''
#     Understanding the processes driving speciation is fundamental to understanding how biodiversity is generated. The two main forces underlying speciation, ecological divergence and sexual selection, are well characterised, yet how they interact during the speciation process is remarkably poorly understood. This project aims to test hypotheses regarding how ecological divergence and sexual selection interact during speciation, from its inception to its completion. In doing so, this research aims to identify genomic regions underlying divergence in colour patterns, which are important for ecological and sexual interactions. Consequently, this project will significantly enhance our understanding of ecological and genetic mechanisms underlying speciation.
#     '''
    with open('stdout.txt','w',encoding="utf-8") as f:
        old = sys.stdout
        sys.stdout = f
        word_list = nltk.word_tokenize(text)
        abstract = nltk.Text(word_list)
        abstract.collocations()
        sys.stdout = old
    
    with open('stdout.txt','r',encoding="utf-8") as f:
        data = f.read()
    
    keywords = list(map(lambda x:x.strip(),data.split(";")))
#     print(keywords)
    return keywords    
    
    
#手动二元搭配
def test_bigramcollocationfinder():
    root_dir = r'E:\github_repo\python_basic\pythonbasictest\self_nltk\files\test.txt'
    sw = stopwords.words("english")
    print(sw)
    with open(root_dir,'r',encoding="utf-8") as f:
        content = f.read()
    content = re.sub('[,;.!-:\(\)“”"\'’‘]','',content)
    word_list = jieba.cut(content)
    word_list = [w for w in word_list if w.strip() and w not in sw and len(w) > 3]
    
    bigram_finder = BigramCollocationFinder.from_words(word_list)
    for (key1,key2),feq in bigram_finder.ngram_fd.items():
        print(key1,key2,feq)
    
#手动三元搭配        
def test_trigramcollocationfinder():
    root_dir = r'E:\github_repo\python_basic\pythonbasictest\self_nltk\files\test.txt'
    sw = stopwords.words("english")
    print(sw)
    with open(root_dir,'r',encoding="utf-8") as f:
        content = f.read()
    content = re.sub('[,;.!-:\(\)“”"\'’‘]','',content)
    word_list = jieba.cut(content)
    word_list = [w for w in word_list if w.strip() and w not in sw and len(w) > 3]
    
    trigram_finder = TrigramCollocationFinder.from_words(word_list)
    
    for (key1,key2,key3),feq in trigram_finder.ngram_fd.items():
        print(key1,key2,key3,feq)
        
#根据词性的三元搭配
def tagged_trigram_collocation(document):
#     content = re.sub('[,;.!-:\(\)“”"\'’‘]','',document)
    
    tag_filter = (
        ('CC','NN','NNS','NNP','NNPS','IN','JJ','JJR','JJS'),
        ('CC','NN','NNS','NNP','NNPS','IN','JJ','JJR','JJS'),
        ('NN','NNS','NNP','NNPS')
    )
    tag_func = lambda key1,key2,key3:key1[1] not in tag_filter[0] or key2[1] not in tag_filter[1] or key3[1] not in tag_filter[2]
    
    #每个单词
    words = nltk.word_tokenize(document)
    
    #每个单词标注词性
    tagged_words = nltk.pos_tag(words)
    
    #全部转换成小写
    tagged_words = ((tw[0].lower(),tw[1]) for tw in tagged_words)
    
    #英语停用词
    sw = stopwords.words("english")
    words = [w for w in tagged_words if w[0].strip() and w[0] not in sw and len(w[0]) > 3]
    
    trigram_finder = TrigramCollocationFinder.from_words(words)
    trigram_finder.apply_ngram_filter(tag_func)
    
    for (key1,key2,key3),feq in trigram_finder.ngram_fd.items():
        print(key1,key2,key3,feq)

#根据词性的二元搭配
def tagged_bigram_collocation(document):
#     content = re.sub('[,;.!-:\(\)“”"\'’‘]','',document)
    
    tag_filter = (
#         ('CC','NN','NNS','NNP','NNPS','IN','JJ','JJR','JJS'),
        ('NN','NNS','NNP','NNPS','JJ','JJR','JJS'),
        ('NN','NNS','NNP','NNPS')
    )
    tag_func = lambda key1,key2:key1[1] not in tag_filter[0] or key2[1] not in tag_filter[1]
    
    #每个单词
    try:
        words = nltk.word_tokenize(document)
    except Exception as e:
        if not isinstance(document,str):
            words = str(document)
        else:
            print(e)
            print(document)
    #每个单词标注词性
    tagged_words = nltk.pos_tag(words)
    
    #全部转换成小写
    tagged_words = ((tw[0].lower(),tw[1]) for tw in tagged_words)
    
    #英语停用词
    sw = stopwords.words("english")
    words = [w for w in tagged_words if w[0].strip() and w[0] not in sw and len(w[0]) > 3]
    
    trigram_finder = BigramCollocationFinder.from_words(words)
    trigram_finder.apply_ngram_filter(tag_func)
    
#     for (key1,key2),feq in trigram_finder.ngram_fd.items():
#         print(key1,key2,feq)
    return trigram_finder


#读取excel的数据    
def get_excel_text():
    root_dir = r'E:\github_repo\python_basic\pythonbasictest\self_nltk\files'
    
    australia = os.path.join(root_dir,'australia.xlsx')
#     america = os.path.join(root_dir,'america.xlsx')
    
    australia_xlsx = load_workbook(australia)
#     america_xlsx = load_workbook(america)
    
    aus_summary_1 = australia_xlsx['Projects']
    aus_summary_2 = australia_xlsx['Fellowships']
    
#     ame_summary_1 = america_xlsx['Sheet1']
#     ame_summary_2 = america_xlsx['Sheet2']
    
    return aus_summary_1,aus_summary_2#,ame_summary_1,ame_summary_2

def config_db(host,username,password,dbname,port=3306):
    db = pymysql.connect(host,username,password,dbname,port)
#     db = pymysql.connect("127.0.0.1","root","","indextest")
    return db


#真实数据测试
def test():
    db = config_db("127.0.0.1","root","","indextest")
    cursor = db.cursor()
    
    print("start ...")
    aus1,aus2 = get_excel_text() #,ame1,ame2
    max_row = aus1.max_row
#     max_column = aus1.max_column
    temp = []
    for r in range(2,max_row + 1):
        if not aus1.cell(row=r,column=2).value:
            continue
        summary = aus1.cell(row=r,column=2).value
        res = tagged_bigram_collocation(summary)
         
        for (key1,key2),feq in res.ngram_fd.items():
            temp.append((key1[0] + " " + key2[0],r,aus1.cell(row=r,column=3).value))
         
    sql = " INSERT INTO `australia_index`(`keyname`,`relatedid`,`year`) VALUES(%s,%s,%s); "
    cursor.executemany(sql,temp)
    db.commit()
    print("the part one end.")
     
    temp = []
    max_row = aus2.max_row
    for r in range(2,max_row + 1):
        if not aus2.cell(row=r,column=2).value:
            continue
        summary = aus2.cell(row=r,column=2).value
        res = tagged_bigram_collocation(summary)
         
        for (key1,key2),feq in res.ngram_fd.items():
            temp.append((key1[0] + " " + key2[0],r,aus1.cell(row=r,column=3).value))
         
    sql = " INSERT INTO `australia_fellowship_index`(`keyname`,`relatedid`,`year`) VALUES(%s,%s,%s); "
    cursor.executemany(sql,temp)
    db.commit()
    print("the part two end.")
    
#     temp = []
#     max_row = ame1.max_row
#     for r in range(1,max_row + 1):
#         if not ame1.cell(row=r,column=2).value:
#             continue
#         summary = ame1.cell(row=r,column=2).value
#         res = tagged_bigram_collocation(summary)
#         
#         for (key1,key2),feq in res.ngram_fd.items():
#             temp.append((key1[0] + " " + key2[0],r))
#         
#     sql = " INSERT INTO `america_index`(`keyword`,`relatedid`) VALUES(%s,%s); "
#     cursor.executemany(sql,temp)
#     db.commit()
#     print("the part three end.")
#     
#     temp = []
#     max_row = ame2.max_row
#     for r in range(1,max_row + 1):
#         if not ame2.cell(row=r,column=2).value:
#             continue
#         summary = ame2.cell(row=r,column=2).value
#         res = tagged_bigram_collocation(summary)
#         
#         for (key1,key2),feq in res.ngram_fd.items():
#             temp.append((key1[0] + " " + key2[0],r))
#         
#     sql = " INSERT INTO `america_index`(`keyword`,`relatedid`) VALUES(%s,%s); "
#     cursor.executemany(sql,temp)
#     db.commit()
#     print("all is ended.")
#     
#     cursor.close()
#     db.close()

def test2():
    db = config_db("127.0.0.1","root","","indextest")
    cursor = db.cursor()
    
    print("start ...")
    aus1,aus2,ame1,ame2 = get_excel_text()

    max_row = aus1.max_row
#     max_column = aus1.max_column
    temp = []
    for r in range(2,max_row + 1):
        if not aus1.cell(row=r,column=2).value:
            continue
        summary = aus1.cell(row=r,column=2).value.lower()
        res = get_phrase_2(summary)

        for key in res:
            if not key:
                continue
#             key1,key2 = key.split()
            temp.append((key,r))
         
    sql = " INSERT INTO `australia_index`(`keyword`,`relatedid`) VALUES(%s,%s); "
    cursor.executemany(sql,temp)
    db.commit()
    print("the part one end.")
     
    temp = []
    max_row = aus2.max_row
    for r in range(2,max_row + 1):
        if not aus2.cell(row=r,column=2).value:
            continue
        summary = aus2.cell(row=r,column=2).value.lower()
        res = get_phrase_2(summary)
         
        for key in res:
            if not key:
                continue
#             key1,key2 = key.split()
            temp.append((key,r))
         
    sql = " INSERT INTO `australia_fellowship_index`(`keyword`,`relatedid`) VALUES(%s,%s); "
    cursor.executemany(sql,temp)
    db.commit()
    print("the part two end.")
    
    temp = []
    max_row = ame1.max_row
    for r in range(1,max_row + 1):
        if not ame1.cell(row=r,column=2).value:
            continue
        summary = ame1.cell(row=r,column=2).value.lower()
        res = get_phrase_2(summary)
        
        for key in res:
            if not key:
                continue
#             key1,key2 = key.split()
            temp.append((key,r))
        
    sql = " INSERT INTO `america_index_copy`(`keyword`,`relatedid`) VALUES(%s,%s); "
    cursor.executemany(sql,temp)
    db.commit()
    print("the part three end.")
    
    temp = []
    max_row = ame2.max_row
    for r in range(1,max_row + 1):
        if not ame2.cell(row=r,column=2).value:
            continue
        summary = ame2.cell(row=r,column=2).value.lower()
        res = get_phrase_2(summary)
        
        for key in res:
            if not key:
                continue
#             key1,key2 = key.split()
            temp.append((key,r))
        
    sql = " INSERT INTO `america_index_copy`(`keyword`,`relatedid`) VALUES(%s,%s); "
    cursor.executemany(sql,temp)
    db.commit()
    print("all is ended.")
    
    cursor.close()
    db.close()





class ConsecutiveNPChunkTagger(nltk.TaggerI):
    def __init__(self, train_sents):
        train_set = []
        for tagged_sent in train_sents:
            untagged_sent = nltk.tag.untag(tagged_sent)
            history = []
            for i, (word, tag) in enumerate(tagged_sent):
                featureset = npchunk_features(untagged_sent, i, history)
                train_set.append( (featureset, tag) )
                history.append(tag)
            self.classifier = nltk.MaxentClassifier.train(train_set, algorithm='megam', trace=0)
            
    def tag(self, sentence):
        history = []
        for i, word in enumerate(sentence):
            featureset = npchunk_features(sentence, i, history)
            tag = self.classifier.classify(featureset)
            history.append(tag)
        return zip(sentence, history)
    
class ConsecutiveNPChunker(nltk.ChunkParserI):
    
    def __init__(self, train_sents):
        tagged_sents = [[((w,t),c) for (w,t,c) in nltk.chunk.tree2conlltags(sent)] for sent in train_sents]
        
        self.tagger = ConsecutiveNPChunkTagger(tagged_sents)
    
    def parse(self, sentence):

        tagged_sents = self.tagger.tag(sentence)
        conlltags = [(w,t,c) for ((w,t),c) in tagged_sents]
        return nltk.chunk.conlltags2tree(conlltags)


#有监督学习的分块器
def tags_since_dt(sentence, i):
    tags = set()
    for word, pos in sentence[:i]:
        if pos == 'DT':
            tags = set()
        else:
            tags.add(pos)
    return '+'.join(sorted(tags))
 
 
 
def npchunk_features(sentence, i, history):
    word, pos = sentence[i]
    if i == 0:
        prevword, prevpos = "<START>", "<START>"
    else:
        prevword, prevpos = sentence[i-1]
    if i == len(sentence)-1:
        nextword, nextpos = "<END>", "<END>"
    else:
        nextword, nextpos = sentence[i+1]
    return {
        "pos": pos,
        "word": word,
        "prevpos": prevpos,
        "nextpos": nextpos,
        "prevpos+pos": "%s+%s" % (prevpos, pos),
        "pos+nextpos": "%s+%s" % (pos, nextpos),
        "tags-since-dt": tags_since_dt(sentence, i)
    }

class UnigramChunker(nltk.ChunkParserI):
    def __init__(self, train_sents): 
        train_data = [[(t,c) for w,t,c in nltk.chunk.tree2conlltags(sent)] for sent in train_sents]
        self.tagger = nltk.UnigramTagger(train_data) 
     
    def parse(self, sentence): 
        pos_tags = [pos for (word,pos) in sentence]
        tagged_pos_tags = self.tagger.tag(pos_tags)
        chunktags = [chunktag for (pos, chunktag) in tagged_pos_tags]
        conlltags = [(word, pos, chunktag) for ((word,pos),chunktag) in zip(sentence, chunktags)]
         
        return nltk.chunk.conlltags2tree(conlltags)



def test3():
    db = config_db("127.0.0.1","root","","indextest")
    cursor = db.cursor()
    
    #分块器训练
    train_sents = conll2000.chunked_sents('train.txt', chunk_types=['NP'])
#     chunker = ConsecutiveNPChunker(train_sents)
    chunker = UnigramChunker(train_sents)
    
    print("start ...")
    aus1,aus2 = get_excel_text() #,ame1,ame2
    max_row = aus1.max_row
    
    
    temp = []
    for r in range(2,max_row + 1):
        if not aus1.cell(row=r,column=2).value:
            continue
        summary = aus1.cell(row=r,column=2).value
        sentences = ie_proprocess(summary)
        
        for st in sentences:
            res = chunker.parse(st)
            for _r in res:
                if isinstance(_r,list):
                    temp.append((" ".join(map(lambda x:x[0],_r)),r,aus1.cell(row=r,column=3).value))
                    
         
    sql = " INSERT INTO `australia_test_index`(`keyname`,`relatedid`,`year`) VALUES(%s,%s,%s); "
    cursor.executemany(sql,temp)
    db.commit()
    print("the part one end.")
     
    temp = []
    for r in range(2,max_row + 1):
        if not aus2.cell(row=r,column=2).value:
            continue
        summary = aus2.cell(row=r,column=2).value
        sentences = ie_proprocess(summary)
        
        for st in sentences:
            res = chunker.parse(st)
            for _r in res:
                if isinstance(_r,list):
                    temp.append((" ".join(map(lambda x:x[0],_r)),r,aus2.cell(row=r,column=3).value))
         
    sql = " INSERT INTO `australia_test_fellow_index`(`keyname`,`relatedid`,`year`) VALUES(%s,%s,%s); "
    cursor.executemany(sql,temp)
    db.commit()
    print("the part two end.")


def regexp():
    sentence = [("the", "DT"), ("little", "JJ"), ("yellow", "JJ"),("dog", "NN"), ("barked", "VBD"), ("at", "IN"), ("the", "DT"), ("cat", "NN")]
    grammar = "NP: {<JJ>*<NN>}"
    cp = nltk.RegexpParser(grammar)
    result = cp.parse(sentence)
    print(result)

if __name__ == "__main__":
#     document = '''
#       This reference manual describes the Python programming language. It is not intended as a tutorial.
# While I am trying to be as precise as possible, I chose to use English rather than formal specifications for everything except syntax and lexical analysis. This should make the document more understandable to the average reader, but will leave room for ambiguities. Consequently, if you were coming from Mars and tried to re-implement Python from this document alone, you might have to guess things and in fact you would probably end up implementing quite a different language. On the other hand, if you are using Python and wonder what the precise rules about a particular area of the language are, you should definitely be able to find them here. If you would like to see a more formal definition of the language, maybe you could volunteer your time — or invent a cloning machine :-).
# It is dangerous to add too many implementation details to a language reference document — the implementation may change, and other implementations of the same language may work differently. On the other hand, CPython is the one Python implementation in widespread use (although alternate implementations continue to gain support), and its particular quirks are sometimes worth being mentioned, especially where the implementation imposes additional limitations. Therefore, you’ll find short “implementation notes” sprinkled throughout the text.
# Every Python implementation comes with a number of built-in and standard modules. These are documented in The Python Standard Library. A few built-in modules are mentioned when they interact in a significant way with the language definition.
# 
#     '''
#     get_phrase()
#     test_bigramcollocationfinder()
#     test_trigramcollocationfinder()
#     tagged_bigram_collocation(document)
    
#     test()
#     get_phrase_2()
#     test()
    
#     regexp()
    test3()
    
    
    
    
    
    
    