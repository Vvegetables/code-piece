from openpyxl import Workbook

#new excel
wb = Workbook()
#active = 0
ws = wb.active
#create sheet
ws1 = wb.create_sheet("title1", index=None)
#modify title
ws.title = "modify_title"
#fill color of sheet title
ws.sheet_properties.tabColor = "1072BA"
#select a sheet
ws1 = wb["title1"]
#loop through worksheets
print(wb.sheetnames)
for sheet in wb:
    print(sheet.title)

#sheet 拷贝
source = wb.active
target = wb.copy_worksheet(source) #有限制

#loop through cells
for row in ws.iter_rows(min_row=1,max_col=3,max_row=2):
    for cell in row:
        print(cell)

for col in ws.iter_cols(min_row=1,max_col=3,max_row=2):
    for cell in col:
        print(cell)

print(tuple(ws.rows))
print(tuple(ws.columns))

#value
for row in ws.values:
    for value in row:
        print(value)
        
        
        
        
        
        
        
